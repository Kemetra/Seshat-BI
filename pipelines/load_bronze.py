"""Load C086 sales Excel files into the bronze layer (raw text + lineage).

Medallion: ensures bronze/silver/gold schemas exist; populates bronze only.
Bronze = faithful landing: every source column as TEXT, plus lineage columns
`_source_file` and `_loaded_at`. No cleaning or typing happens here — that is
the silver layer's job. Landing raw text keeps the load reproducible and lets
any value be traced back to its source file.

Idempotent: a file's prior rows are deleted before it is re-loaded, so a re-run
replaces rather than duplicates. Loads via PostgreSQL COPY (fast over TLS).
Each file is reconciled (Excel data rows == bronze rows) before moving on.

Secrets are never hardcoded. The DB connection (host/port/user/password) is read
from `doctl` at runtime. Provide the cluster id and source dir via flags or env.

Usage:
    python load_bronze.py --create-only
    python load_bronze.py --file C086_Sales_2024_Full_Year.xlsx
    python load_bronze.py --all

Configuration (flags override env):
    --cluster-id / DO_DB_CLUSTER_ID   DigitalOcean DB cluster id (for doctl)
    --database   / ANALYTICS_DB_NAME  target logical database (no default; supply via env/flag)
    --src-dir    / SALES_SRC_DIR      folder of .xlsx files to load
    --base-file                       file whose header defines the bronze columns
"""

from __future__ import annotations

import argparse
import csv
import datetime
import io
import json
import os
import re
import subprocess
import sys

import openpyxl
import psycopg2

BRONZE_TABLE = "bronze.sales_c086_raw"


def log(msg: str) -> None:
    print(f"[{datetime.datetime.now():%H:%M:%S}] {msg}", flush=True)


def get_conn_params(cluster_id: str, database: str) -> dict:
    """Read connection details from doctl at runtime (no secrets in source)."""
    out = subprocess.run(
        ["doctl", "databases", "connection", cluster_id, "-o", "json"],
        capture_output=True,
        text=True,
        check=True,
    )
    info = json.loads(out.stdout)
    c = info[0] if isinstance(info, list) else info
    return {
        "host": c["host"],
        "port": int(c["port"]),
        "user": c["user"],
        "password": c["password"],
        "dbname": database,
        "sslmode": "require",
    }


def connect(cluster_id: str, database: str):
    conn = psycopg2.connect(**get_conn_params(cluster_id, database))
    conn.autocommit = True
    with conn.cursor() as cur:
        cur.execute("SET client_encoding TO 'UTF8';")
    return conn


def norm_headers(headers: list[str]) -> list[str]:
    """Deterministic snake_case; collision-safe (duplicate names get _N suffix)."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in headers:
        s = re.sub(r"[^\w]+", "_", (h or "").strip().lower(), flags=re.UNICODE)
        s = re.sub(r"_+", "_", s).strip("_") or "col"
        if s in seen:
            seen[s] += 1
            s = f"{s}_{seen[s]}"
        else:
            seen[s] = 1
        out.append(s)
    return out


def read_base_headers(src_dir: str, base_file: str) -> list[str]:
    wb = openpyxl.load_workbook(
        os.path.join(src_dir, base_file), read_only=True, data_only=True
    )
    ws = wb[wb.sheetnames[0]]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    return norm_headers(hdr)


def create_objects(conn, cols: list[str]) -> None:
    with conn.cursor() as cur:
        for sch in ("bronze", "silver", "gold"):
            cur.execute(f"CREATE SCHEMA IF NOT EXISTS {sch};")
        coldefs = ",\n  ".join(f'"{c}" TEXT' for c in cols)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {BRONZE_TABLE} (
              {coldefs},
              _source_file TEXT NOT NULL,
              _loaded_at   TIMESTAMPTZ NOT NULL DEFAULT now()
            );
        """)
    log(
        f"schemas bronze/silver/gold ready; {BRONZE_TABLE} = {len(cols)} cols + lineage"
    )


def excel_rows(path: str):
    """Yield each data row as faithful text (None -> '')."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        yield ["" if v is None else str(v) for v in row]
    wb.close()


def csv_buffer(path: str, ncols: int, source_file: str) -> tuple[io.StringIO, int]:
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_ALL, lineterminator="\n")
    loaded_at = datetime.datetime.now(datetime.timezone.utc).isoformat()
    n = 0
    for r in excel_rows(path):
        r = (list(r) + [""] * ncols)[:ncols]  # pad/truncate defensively
        w.writerow(r + [source_file, loaded_at])
        n += 1
    buf.seek(0)
    return buf, n


def load_file(conn, cols: list[str], src_dir: str, fname: str) -> None:
    path = os.path.join(src_dir, fname)
    if not os.path.exists(path):
        sys.exit(f"file not found: {path}")
    quoted = ", ".join(f'"{c}"' for c in cols) + ', "_source_file", "_loaded_at"'
    with conn.cursor() as cur:
        cur.execute(f"DELETE FROM {BRONZE_TABLE} WHERE _source_file = %s;", (fname,))
        log(f"{fname}: cleared prior rows; reading + copying...")
        buf, n = csv_buffer(path, len(cols), fname)
        cur.copy_expert(
            f"COPY {BRONZE_TABLE} ({quoted}) FROM STDIN "
            f"WITH (FORMAT csv, QUOTE '\"', ENCODING 'UTF8')",
            buf,
        )
        log(f"{fname}: COPY done, {n} rows sent")
    reconcile(conn, fname, path)


def reconcile(conn, fname: str, path: str) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    excel_rows_n = wb[wb.sheetnames[0]].max_row - 1
    wb.close()
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT COUNT(*) FROM {BRONZE_TABLE} WHERE _source_file = %s;", (fname,)
        )
        db_rows = cur.fetchone()[0]
    ok = excel_rows_n == db_rows
    verdict = "PASS" if ok else "FAIL"
    log(f"RECONCILE {fname}: excel={excel_rows_n} db={db_rows} -> {verdict}")
    if not ok:
        sys.exit(f"reconcile FAILED for {fname}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Load C086 sales Excel files into bronze.")
    ap.add_argument("--cluster-id", default=os.environ.get("DO_DB_CLUSTER_ID"))
    ap.add_argument(
        "--database", default=os.environ.get("ANALYTICS_DB_NAME")
    )
    ap.add_argument("--src-dir", default=os.environ.get("SALES_SRC_DIR"))
    ap.add_argument("--base-file", default="C086_Sales_2023_H1_Jan-Jun.xlsx")
    ap.add_argument("--create-only", action="store_true")
    ap.add_argument("--file")
    ap.add_argument("--all", action="store_true")
    a = ap.parse_args()

    if not a.cluster_id:
        sys.exit("missing --cluster-id (or DO_DB_CLUSTER_ID env)")
    if not a.create_only and not a.src_dir:
        sys.exit("missing --src-dir (or SALES_SRC_DIR env)")

    conn = connect(a.cluster_id, a.database)
    log(f"connected to {a.database}")
    cols = read_base_headers(a.src_dir, a.base_file) if a.src_dir else None
    if cols is None:
        sys.exit("--create-only still needs --src-dir to read the column header")
    create_objects(conn, cols)

    if a.file:
        load_file(conn, cols, a.src_dir, a.file)
    elif a.all:
        files = sorted(f for f in os.listdir(a.src_dir) if f.lower().endswith(".xlsx"))
        for f in files:
            load_file(conn, cols, a.src_dir, f)
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM {BRONZE_TABLE};")
            log(f"TOTAL bronze rows: {cur.fetchone()[0]}")
    conn.close()
    log("done")


if __name__ == "__main__":
    main()
